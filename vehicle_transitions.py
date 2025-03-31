import openpyxl
import vehicle_options
from vehicle_options import Vehicle

# Establish Excel data
path = "C:\\Users\\cejva\\Documents\\Work\\FM Fleet Vehicles from Interviews_FINAL.xlsx"
workbook = openpyxl.load_workbook(path)
sheet = workbook['raw_input']

# Sheet column values
department = 1
VID = 3
year = 4
make_model = 5
awk_material = 8
weight = 9
size = 10
passengers = 19
add_passengers = 20
sharing = 21
outer_sharing = 22
outer_tasks = 23
tasks = 24
scheduled = 25
usage_percent = 26
drive_dist = 28
v_mods = 29

# Misc. Variables


# Define vehicle rec data objects
lightning = Vehicle(name="Ford F150 Lightning", sizes=["Medium", "Large"], weights=["Heavy", "Heavy Duty"], dists=["Any"], scheduling=["Any"], share_tasks="Any", use_low=0.60, use_high=1.0)
box = Vehicle(name="Peterbilt 220EV", sizes=["Heavy Duty"], weights=["Medium", "Heavy"], dists=["Any"], scheduling=["Scheduled", "Possible"], share_tasks="Any", use_low=0.0, use_high=0.60)
transit = Vehicle(name="Ford eTransit", sizes=["Medium", "Large"], weights=["Medium", "Heavy", "Heavy Duty"], dists=["Any"], scheduling=["Unscheduled", "Possible"], share_tasks="No", use_low=0.60, use_high=1.0)
moto_truck = Vehicle(name="MotoEV Utility Truck", sizes=["Medium", "Large"], weights=["Medium", "Heavy", "Heavy Duty"], dists=["Campus-bound", "Between campuses"], scheduling=["Any"], share_tasks="Any", use_low=0.60, use_high=1.0)
utv = Vehicle(name="MotoEV UTV", sizes=["Small", "Medium"], weights=["Light", "Medium", "Heavy"], dists=["Campus-bound", "Between campuses"], scheduling=["Any"], share_tasks="Any", use_low=0, use_high=1.0)
trike = Vehicle(name="Civilized Cycles Semi-Trike", sizes=["Small", "Medium", "Large"], weights=["Light", "Medium"], dists=["Campus-bound"], scheduling=["Unscheduled", "Possible"], share_tasks="Any", use_low=0, use_high=1.0)

for x in range(2, 193):
    ### variables
    # Vehicle Options to recommend at end step
    options = [trike, utv, moto_truck, box, transit, lightning] # in ascending order of additional infrastructure requirement
    # Vehicles to remove from final recommendation list
    removals = []
    shareable = 0
    
    # sheet data
    usage = float(sheet.cell(x, usage_percent).value)
    out_sharing = sheet.cell(x, outer_sharing).value
    task_share = sheet.cell(x, outer_tasks).value
    sched = sheet.cell(x, scheduled).value
    curr_dist = sheet.cell(x, drive_dist).value
    curr_size = sheet.cell(x, size).value
    curr_weight = sheet.cell(x, weight).value
    
    # determine shareability/sharepool potential (0-3)
    if sched != "Unscheduled":
        shareable += 1
    if task_share != "No":
        shareable += 1
    if out_sharing != "No":
        shareable += 1

    # Peterbilt requires some level of shareability
    if shareable < 1:
        removals.append(box.name)
    #print(sheet.cell(x, make_model).value)
    # refine vehicle options
    for v in options:
        #print(usage)
        #print(v.use_high)
        #print(usage > v.use_high)
        # usage
        if (v.use_low > usage) or (usage > v.use_high):
            removals.append(v.name)
        # task sharing
        elif (task_share not in v.share_tasks) and (v.share_tasks != "Any"):
            removals.append(v.name)
        # scheduling
        elif (sched not in v.scheduling) and ("Any" not in v.scheduling):
            removals.append(v.name)
        # distances
        elif (curr_dist not in v.dists) and ("Any" not in v.dists):
            removals.append(v.name)
        # size
        elif curr_size not in v.sizes:
            removals.append(v.name)
        # weight
        elif curr_weight not in v.weights:
            removals.append(v.name)

    # remove unfit recommendations
    for v in options:
        if removals.count(v.name) > 0:
            options.remove(v)

    # write selection(s) to sheet
    init_col = 32
    if options.__sizeof__() == 1:
        sheet.cell(x, 32).value = options[0].name
        init_col += 1
    else:
        for v in options:
            print(v.name)
            sheet.cell(x, init_col).value = v.name
            init_col += 1
    # write vehicle shareability score to sheet
    sheet.cell(x, init_col).value = shareable

    # SAVE
    #workbook.save(path)
        

'''
for each row
    usage
        cycle each curr option - remove those with ranges not including
        if peterbilt range includes && NOT unscheduled
            sharepool flag (still recc smaller vehicle
    sizes
        
    weights
    
    
'''